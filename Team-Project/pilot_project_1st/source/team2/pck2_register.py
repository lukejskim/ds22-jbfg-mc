import pck1_common
import json
# import openpyxl
import random
from openpyxl.utils import get_column_letter  # 컬럼값을 숫자값으로 변경


# 미션1 : 참가선수들 등록
def reg_player(excel_file):
    # 참가자 명단 json파일로 준비
    with open('./data/golf_players.json', 'w', encoding='UTF8') as fp:
        data ='["김영목","박성실","박요온","한혜형","오승현","김은민","박동현","이건호","정은지","최대훈"]'
        fp.write(data)

    is_success = False
    result_msg = "미션1_reg_player"

    wb, ws = pck1_common.excel_loading(excel_file)

    with open('./data/golf_players.json', 'r', encoding='UTF8') as data_file:
        golf_players = json.load(data_file)

    for r_no in range(3, 13):
        player_loc = "{}{}".format('A', r_no)
        ws[player_loc] = golf_players[r_no - 3]  # 리스트 처음 이름부터 차례대로
        # print(golf_players[r_no-3])

    #
    # ws['A3'] = '김영목'
    # ws['A4'] = '박성실'
    # ws['A5'] = '박요온'
    # ws['A6'] = '한혜형'
    # ws['A7'] = '오승현'
    # ws['A8'] = '김은민'
    # ws['A9'] = '박동현'
    # ws['A10'] = '이건호'
    # ws['A11'] = '정은지'
    # ws['A12'] = '최대훈'

    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg

# 미션2 : 참가선수들 HOLE별 스코어를 랜덤하게 입력 (-2 ~ PAR)
def reg_score_mgr(excel_file):
    is_success = False
    result_msg = "미션2_reg_score"

    wb, ws = pck1_common.excel_loading(excel_file)

    for c_no in range(3, 21):
        for r_no in range(3, 13):
            c_nm = get_column_letter(c_no)
            par_row = 2
            par_loc = "{}{}".format(c_nm, par_row)
            par_val = ws[par_loc].value
            random_number = random.randint(-2, par_val)
            loc = "{}{}".format(get_column_letter(c_no), r_no)
            ws[loc].value = random_number

        # msg = "{} : ws[{}].value = {}".format(c_no, loc, random_number)
        # print(msg)
    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg


# For Module Test!!!!
if __name__ == "__main__":
    # excel_file = './data/golf_score_board_test.xlsx'
    # is_success = reg_score_mgr(excel_file)
    # print(is_success)
    print('random_scoring_main....')

    # wb.save(excel_file)
#     wb.close()