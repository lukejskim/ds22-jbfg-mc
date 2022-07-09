import pck1_common
# import json
# import openpyxl
# import random
from openpyxl.utils import get_column_letter  # 컬럼값을 숫자값으로 변경


# 미션3 : 18홀까지의 스코어, 보정치를 계산하여 등록 (핸디는 옵션)

## (1) 스트로크 계산하기
def cal_stroke(excel_file):
    is_success = False
    result_msg = "미션3(1)_stroke"

    wb, ws = pck1_common.excel_loading(excel_file)

    for r_no in range(3, 13):
        sum_strk_val = int(72)
        for c_no in range(3, 21):
            c_nm = get_column_letter(c_no)
            strk_data_loc = "{}{}".format(c_nm, r_no) #r_no
            strk_data_val = ws[strk_data_loc].value
            sum_strk_val = sum_strk_val + strk_data_val
            strk_loc = "{}{}".format('U', r_no) # r_no
            ws[strk_loc].value = sum_strk_val
    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg

## (2) 보정치 Default=0
def cal_handy(excel_file):
    is_success = False
    result_msg = "미션3(2)_handy"

    wb, ws = pck1_common.excel_loading(excel_file)

    for r_no in range(3, 13):
        handy_val = 0
        handy_loc = "{}{}".format('v', r_no)  # r_no
        ws[handy_loc].value = handy_val
    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg

## (3) 최종성적
def cal_final_score(excel_file):
    is_success = False
    result_msg = "미션3(3)_final_score"

    wb, ws = pck1_common.excel_loading(excel_file)

    for r_no in range(3, 13):
        final_strk = "{}{}".format('U', r_no)
        final_handy = "{}{}".format('V', r_no)
        final_score = ws[final_strk].value - ws[final_handy].value
        final_loc = "{}{}".format('w', r_no)
        ws[final_loc].value = final_score  # 최종성적

    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg



# For Module Test!!!!
if __name__ == "__main__":
    # excel_file = './data/golf_score_board_test.xlsx'
    # is_success = reg_score_mgr(excel_file)
    # print(is_success)
    print('random_scoring_main....')

    # wb.save(excel_file)
#     wb.close()