import pck1_common
# import json
# import openpyxl
# import random
from openpyxl.utils import get_column_letter  # 컬럼값을 숫자값으로 변경

## (4) ranking
def cal_ranking(excel_file):
    is_success = False
    result_msg = "미션4(1)_ranking"

    wb, ws = pck1_common.excel_loading(excel_file)

    final_score_list = list()
    for r_no in range(3, 13):
        final_loc = "{}{}".format('w', r_no)
        final_score_list.append(ws[final_loc].value)  # 최종성적
        rank_score = sorted(final_score_list)

    for r_no in range(3, 13):
        final_loc = "{}{}".format('w', r_no)  # w5(97)
        for i in range(len(final_score_list)):
            if ws[final_loc].value == rank_score[i]:  # w5(97)값이 7번째랑 같음
                rank_no = i + 1
                rank_loc = "{}{}".format('x', r_no)
                ws[rank_loc].value = rank_no  # 최종랭킹
                break
    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg


# 미션4 : 각종 기록 갯수 등록
def cnt_records(excel_file):
    is_success = False
    result_msg = "미션4(2)_cnt_records"

    wb, ws = pck1_common.excel_loading(excel_file)

    for r_no in range(3, 13):
        eagle = 0
        birdy = 0
        par = 0
        boggy = 0
        par_par = 0

        for c_no in range(3, 21):
            c_nm = get_column_letter(c_no)
            par_loc = "{}{}".format(c_nm, r_no)
            par_val = ws[par_loc].value
            hole_par_loc = "{}{}".format(c_nm, 2)
            hole_par_val = ws[hole_par_loc].value

            if par_val == -2:
                eagle += 1
            elif par_val == -1:
                birdy += 1
            elif par_val == 0:
                par += 1
            elif par_val == 1:
                boggy += 1
            elif par_val == hole_par_val:
                par_par += 1

        eagle_loc = "{}{}".format('Y', r_no)
        ws[eagle_loc].value = eagle
        birdy_loc = "{}{}".format('Z', r_no)
        ws[birdy_loc].value = birdy
        par_loc = "{}{}".format('AA', r_no)
        ws[par_loc].value = par
        boggy_loc = "{}{}".format('AB', r_no)
        ws[boggy_loc].value = boggy
        par_par_loc = "{}{}".format('AC', r_no)
        ws[par_par_loc].value = par_par

    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg