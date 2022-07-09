import pck1_common
# import pck5_result_common as result_common
import json
import openpyxl
import random
from openpyxl.utils import get_column_letter  # 컬럼값을 숫자값으로 변경


# # 미션5 : 대회결과 시트에 결과에 맞는 선수명과 최종성적/기록수를 등록




# 미션5 집계를 위한 선수별 dic정보 리스트 만들기
def result_records(excel_file):
    is_success = False
    result_msg = "미션5(2)_result_records"

    wb, ws = pck1_common.excel_loading(excel_file)

    golfs = list()
    players_name = list()
    ranks_no = list()
    finals_score = list()
    eagles_cnt = list()
    birdys_cnt = list()
    pars_cnt = list()
    boggys_cnt = list()
    parpars_cnt = list()

    for r_no in range(3, 13):
        player_name_loc = "{}{}".format('A', r_no)
        player_name = ws[player_name_loc].value
        rank_no_loc = "{}{}".format('X', r_no)
        rank_no = ws[rank_no_loc].value
        final_score_loc = "{}{}".format('W', r_no)
        final_score = ws[final_score_loc].value
        eagle_cnt_loc = "{}{}".format('Y', r_no)
        eagle_cnt = ws[eagle_cnt_loc].value
        birdy_cnt_loc = "{}{}".format('Z', r_no)
        birdy_cnt = ws[birdy_cnt_loc].value
        par_cnt_loc = "{}{}".format('AA', r_no)
        par_cnt = ws[par_cnt_loc].value
        boggy_cnt_loc = "{}{}".format('AB', r_no)
        boggy_cnt = ws[boggy_cnt_loc].value
        pparpar_cnt_loc = "{}{}".format('AC', r_no)
        parpar_cnt = ws[pparpar_cnt_loc].value

        players_name.append((player_name))
        ranks_no.append((rank_no))
        finals_score.append((final_score))
        eagles_cnt.append((eagle_cnt))
        birdys_cnt.append((birdy_cnt))
        pars_cnt.append((par_cnt))
        boggys_cnt.append((boggy_cnt))
        parpars_cnt.append((parpar_cnt))

    for i in range(10):
        golf = {'이름': players_name[i], '랭킹': ranks_no[i], '최종성적': finals_score[i], '이글': eagles_cnt[i],
                '버디': birdys_cnt[i], '파': pars_cnt[i], '보기': boggys_cnt[i], '양파': parpars_cnt[i]}
        golfs.append(golf)

    wb, ws = pck1_common.excel_loading(excel_file, "대회결과")
    # ws = wb["대회결과"]

    # 대회 결과 입력 빌드업

    # 대회결과 / 기록 / 선수명
    player11 = list()
    player22 = list()
    player33 = list()
    player44 = list()

    # 대회결과 꼴등 / 기록 최댓값을 위한
    real_birdys_cnt = sorted(birdys_cnt)
    real_pars_cnt = sorted(pars_cnt)
    real_boggys_cnt = sorted(boggys_cnt)
    real_parpars_cnt = sorted(parpars_cnt)

    # 대회결과 / 기록 / 선수명
    for golf in golfs:
        if golf['버디'] == real_birdys_cnt[-1]:
            player11.append(golf['이름'])
        if golf['파'] == real_pars_cnt[-1]:
            player22.append(golf['이름'])
        if golf['보기'] == real_boggys_cnt[-1]:
            player33.append(golf['이름'])
        if golf['양파'] == real_parpars_cnt[-1]:
            player44.append(golf['이름'])

    names = ''
    for name in player11:
        if len(names) == 0:
            names = name
        else:
            names += ', ' + name
    player11 = names
    ws["B8"] = player11

    names = ''
    for name in player22:
        if len(names) == 0:
            names = name
        else:
            names += ', ' + name
    player22 = names
    ws["B9"] = player22

    names = ''
    for name in player33:
        if len(names) == 0:
            names = name
        else:
            names += ', ' + name
    player33 = names
    ws["B10"] = player33

    names = ''
    for name in player44:
        if len(names) == 0:
            names = name
        else:
            names += ', ' + name
    player44 = names
    ws["B11"] = player44

    # 대회결과 / 기록 / 기록수
    for golf in golfs:
        if golf['버디'] == real_birdys_cnt[-1]:
            ws["C8"] = real_birdys_cnt[-1]
        if golf['파'] == real_pars_cnt[-1]:
            ws["C9"] = real_pars_cnt[-1]
        if golf['보기'] == real_boggys_cnt[-1]:
            ws["C10"] = real_boggys_cnt[-1]
        if golf['양파'] == real_parpars_cnt[-1]:
            ws["C11"] = real_parpars_cnt[-1]

    else:
        is_success = True

    if is_success:
        wb.save(excel_file)
        is_success = True
        result_msg += " : 완료"
    else:
        is_success = False
        result_msg += " : 실패"

    return is_success, result_msg

# For Module Test!!!!
if __name__ == "__main__":
    # excel_file = './data/golf_score_board_test.xlsx'
    # is_success = reg_score_mgr(excel_file)
    # print(is_success)
    print('random_scoring_main....')
