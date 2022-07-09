import openpyxl


def regUserScores(randomScore, player_list):
    excel_file = './data/golf_score_board.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["스코어"]

    reg_randomScore = randomScore

    for c in range(0,18):
        for r in range(len(player_list)):
            ws.cell(row=r+3, column=c+3, value=reg_randomScore[r][c])

    wb.save(excel_file)
    return reg_randomScore


def regStroke(player_list, strokeScore):
    excel_file = './data/golf_score_board.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["스코어"]

    for r in range(len(player_list)):
        ws.cell(r+3, column=21, value=strokeScore[r])

    wb.save(excel_file)
    return strokeScore


def regHandy(player_list, handyScore):
    excel_file = './data/golf_score_board.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["스코어"]

    for r in range(len(player_list)):
        ws.cell(r+3, column=22, value=handyScore[r])

    wb.save(excel_file)
    return handyScore


def regTotalScore(player_list,score):
    excel_file = './data/golf_score_board.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["스코어"]

    for r in range(len(player_list)):
        ws.cell(r + 3, column=23, value=score[r])  # 참가자 최종 점수 산출

    wb.save(excel_file)
    return score


def regRank(rank_list):
    excel_file = './data/golf_score_board.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["스코어"]

    for r in range(len(rank_list)):
        ws.cell(r + 3, column=24, value=rank_list[r]) # 참가자 랭크 산출

    wb.save(excel_file)
    return rank_list


def regEbpbCnt(player_list, ebpbCnt):
    excel_file = './data/golf_score_board.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["스코어"]

    for r in range(len(player_list)):
        ws.cell(row=r + 3, column=25, value=ebpbCnt[r][0])
        ws.cell(row=r + 3, column=26, value=ebpbCnt[r][1])
        ws.cell(row=r + 3, column=27, value=ebpbCnt[r][2])
        ws.cell(row=r + 3, column=28, value=ebpbCnt[r][3])

    wb.save(excel_file)
    return ebpbCnt