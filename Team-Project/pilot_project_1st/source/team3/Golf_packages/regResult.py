import openpyxl

results         = []
tmp_playerName  = []
playerName      = []
tmp_playerScore = []
playerScore     = []
boggyList       = []
parList         = []
buddyList       = []
doubleParList   = []
rankList        = []
scoreList       = []


def Result(player_list, score, rank_list, ebpbCnt):
    excel_file = '../Golf/data/golf_score_board.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["대회결과"]

    for i in range(len(player_list)):
        player_result = {"이름" : player_list[i], "최종점수" : score[i], "순위" : rank_list[i], "이글" : ebpbCnt[i][0], "버디" : ebpbCnt[i][1] , "파" : ebpbCnt[i][2]
                        , "보기" : ebpbCnt[i][3], "양파" : ebpbCnt[i][4]}
        results.append(player_result)

    max_rank = max(rank_list)

    row = [1,2,3,max_rank]
    for r in row:
        tmp_playerName.append([result['이름'] for result in results if result['순위'] == r])
        tmp_playerScore.append([result['최종점수'] for result in results if result['순위'] == r])

    playerName = [data for inner_list in tmp_playerName for data in inner_list]
    playerScore = [data for inner_list in tmp_playerScore for data in inner_list]


    for r in range(3):
        ws.cell(r+2, column=2, value=playerName[r])
        ws.cell(r+2, column=3, value=playerScore[r])

    ws.cell(5, column=2, value=playerName[-1])
    ws.cell(5, column=3, value=playerScore[-1])


    for r in range(len(player_list)):
        buddyList.append(results[r]["버디"])
        parList.append(results[r]["파"])
        boggyList.append(results[r]["보기"])
        doubleParList.append(results[r]["양파"])

    buddyIdx = buddyList.index(max(buddyList))
    parIdx = parList.index(max(parList))
    boggyIdx = boggyList.index(max(boggyList))
    doubleParIdx = doubleParList.index(max(doubleParList))

    ws.cell(8, column=2, value=results[buddyIdx]["이름"])
    ws.cell(8, column=3, value=results[buddyIdx]['버디'])
    ws.cell(9, column=2, value=results[parIdx]['이름'])
    ws.cell(9, column=3, value=results[parIdx]['파'])
    ws.cell(10, column=2, value=results[boggyIdx]['이름'])
    ws.cell(10, column=3, value=results[boggyIdx]['보기'])
    ws.cell(11, column=2, value=results[doubleParIdx]['이름'])
    ws.cell(11, column=3, value=results[doubleParIdx]['양파'])

    wb.save(excel_file)
    return results
